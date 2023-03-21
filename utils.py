import base64
import requests
import json
from openpyxl import Workbook, load_workbook


# answer = "25) every village or group of villages has a gram panchayat which consists of severall ward members c panch_] and a presstr(id)ent [sarpanch]. it is elected by all the adult ward members living in that ward it is under the supervision on of the gram sabha which consists of the adult members of that population who meet a few times a year to discuss the performance of the gram sabha and the annual budget. a few gram sabhas form a panchayat samiti or block or mandal elected by all the panchayat members in that area a few panchayat samitis constitute the zilla parishad which 13 consists of mlas members of legislative assemblies, and other ministers. zilla parishad chairperson is the political head of zilla parishad"


def inputDataExcel(answer, id):
    countWords = len(answer)
    wb = load_workbook("Answers-Report.xlsx")
    ws = wb.active
    ws["E" + str(id)].value = answer
    keywords = ws["D" + str(id)].value
    keywords = keywords.split(",")

    if countWords >= 100:
        ws["F" + str(id)].value = 5
    else:
        ws["F" + str(id)].value = 0

    count = 0
    for i in keywords:
        if i in answer:
            count += 1
            if 1 <= count < 3:
                ws["G" + str(id)].value = 2
            elif count >= 3:
                ws["G" + str(id)].value = 5
        else:
            ws["G" + str(id)].value = 0

    totalMarks = ws["F" + str(id)].value + ws["G" + str(id)].value
    ws["H" + str(id)].value = totalMarks

    wb.save("Answers-Report.xlsx")


# inputDataExcel(answer=answer, id=answer[:1])
